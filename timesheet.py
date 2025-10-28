import os
import pandas as pd
from datetime import datetime, timedelta
import sounddevice as sd
from scipy.io.wavfile import write
import openai
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import time
import json
import sys
import argparse
import numpy as np
import soundfile as sf

print("Starting timesheet application...")
print("All modules imported successfully")
print(f"Python version: {sys.version}")

# ==========================
# FILE PATH
# ==========================
FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "timesheet_updated.xlsx")
AUDIO_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "temp_audio.wav")

# Print absolute paths for debugging
print(f"Excel file will be saved to: {FILE_PATH}")
print(f"Audio file will be saved to: {AUDIO_FILE}")
print(f"Current working directory: {os.getcwd()}")

# Try to import whisper more defensively
whisper = None
try:
    import whisper
    print("Whisper module imported successfully")
except ImportError:
    print("Warning: Whisper module not available. Will use OpenAI Whisper API for transcription.")

# Check audio recording modules
try:
    import sounddevice as sd
    from scipy.io.wavfile import write
    print("Audio recording modules imported successfully")
except ImportError:
    print("Warning: Audio recording modules not available. Interactive mode will not work.")
    sd = None

openai.api_key = ""  

def record_audio(file_path, duration=15, sample_rate=16000):
    """Record audio for a specified duration and save to file_path"""
    try:
        # Ensure directory exists
        directory = os.path.dirname(os.path.abspath(file_path))
        if directory:
            os.makedirs(directory, exist_ok=True)
            print(f"Directory exists: {directory}")
        
        print(f"Recording for {duration} seconds...")
        
        # Countdown
        for i in range(3, 0, -1):
            print(f"Starting in {i}...")
            time.sleep(1)
        
        # Record audio
        print("Recording NOW - Speak clearly...")
        recording = sd.rec(int(duration * sample_rate), samplerate=sample_rate, channels=1, dtype='float32')
        
        # Show progress during recording
        for i in range(duration):
            print(f"Recording: {i+1}/{duration} seconds", end="\r", flush=True)
            time.sleep(1)
        print("\nFinishing recording...")
        
        sd.wait()  # Wait until recording is finished
        
        # Check if recording contains data
        if recording.size == 0:
            print(" Warning: Recording appears to be empty.")
            return False
            
        max_amplitude = np.max(np.abs(recording))
        print(f"Maximum audio level: {max_amplitude:.4f}")
        
        if max_amplitude < 0.01:
            print(" Audio level is very low. Please speak louder or check your microphone.")
            # Still try to save it
        
        # Normalize audio (improve volume)
        if max_amplitude > 0:
            normalized_recording = recording / max_amplitude * 0.9
        else:
            normalized_recording = recording
        
        # Save recording to file using scipy.io.wavfile.write
        print(f"Saving recording to {file_path}...")
        write(file_path, sample_rate, normalized_recording)
        
        # Verify file was created
        if os.path.exists(file_path):
            file_size = os.path.getsize(file_path)
            print(f"Recording saved successfully to {file_path} (Size: {file_size} bytes)")
            return True
        else:
            print(f" Failed to save recording to {file_path}")
            return False
            
    except Exception as e:
        print(f"Error recording audio: {e}")
        import traceback
        traceback.print_exc()
        return False

# ==========================
# TRANSCRIPTION FUNCTIONS
# ==========================
def transcribe_audio(file_path):
    """Transcribe audio file to text using Whisper (locally or API)"""
    print(f"Transcribing audio file: {file_path}")
    
    try:
        # Use local Whisper model if available (preferred for privacy & cost)
        if whisper:
            print("Using local Whisper model...")
            model = whisper.load_model("small")  # Options: tiny, base, small, medium, large
            
            # Load audio file
            audio, sample_rate = load_audio_file(file_path)
            if audio is None:
                raise Exception("Failed to load audio file")
                
            # Transcribe
            result = model.transcribe(file_path)
            transcribed_text = result["text"]
            
            print(f" Transcription successful: {transcribed_text[:100]}...")
            return transcribed_text
        else:
            # Use OpenAI API for transcription
            print("Using OpenAI API for transcription...")
            
            with open(file_path, "rb") as audio_file:
                response = openai.Audio.transcribe(
                    model="whisper-1", 
                    file=audio_file
                )
                
            transcribed_text = response.get("text", "")
            
            print(f"Transcription successful: {transcribed_text[:100]}...")
            return transcribed_text
            
    except Exception as e:
        print(f"Error transcribing audio: {e}")
        import traceback
        traceback.print_exc()
        return None

# ==========================
# GPT PROCESSING
# ==========================
def process_with_gpt(text):
    """Process transcribed text using GPT to extract structured data"""
    try:
        print("Processing with GPT...")
        
        prompt = f"""The following is a transcription of daily activities. Extract the data in the following format:

Date: [MM-DD-YY]  
Day: [Day of the week]  
Start Time: [HH:MM AM/PM]  
End Time: [HH:MM AM/PM]  
Time Elapsed: [X hrs]  
Task: [Task Description]

If a date is not specified, assume today's date. If a day is not specified, derive it from the date.
If multiple activities are mentioned, extract each as a separate entry.

Transcription: '{text}'
"""

        response = openai.ChatCompletion.create(
            model="gpt-4",
            messages=[{"role": "system", "content": "You extract timesheet entries from transcribed text."},
                     {"role": "user", "content": prompt}],
            temperature=0.1
        )
        
        extracted_data = response.choices[0].message.content
        print(f"GPT processing successful")
        print(f"Extracted data:\n{extracted_data}")
        
        return extracted_data
        
    except Exception as e:
        print(f"❌ Error processing with GPT: {e}")
        import traceback
        traceback.print_exc()
        return None

# ==========================
# TIME CALCULATION
# ==========================
def calculate_time_elapsed(start_time_str, end_time_str):
    """Calculate time elapsed between start and end times"""
    try:
        # Parse times
        time_format = "%I:%M %p"
        start_time = datetime.strptime(start_time_str, time_format)
        end_time = datetime.strptime(end_time_str, time_format)
        
        # Handle time span across midnight
        if end_time < start_time:
            end_time += timedelta(days=1)
        
        # Calculate difference
        time_diff = end_time - start_time
        total_minutes = time_diff.total_seconds() / 60
        hours = total_minutes // 60
        minutes = total_minutes % 60
        
        # Format output
        if minutes == 0:
            return f"{int(hours)} hrs"
        else:
            return f"{int(hours)} hrs {int(minutes)} mins"
        
    except Exception as e:
        print(f"Error calculating time elapsed: {e}")
        return "N/A"

# ==========================
# EXCEL FUNCTIONS
# ==========================
def update_excel(date, day, start_time, end_time, time_elapsed, task):
    """Update Excel file with new timesheet entry"""
    try:
        # Create file if it doesn't exist
        if not os.path.exists(FILE_PATH):
            df = pd.DataFrame(columns=["Date", "Day", "Start Time", "End Time", "Time Elapsed", "Task"])
            df.to_excel(FILE_PATH, index=False)
        
        # Load workbook
        wb = load_workbook(FILE_PATH)
        ws = wb.active
        
        # Find the next empty row
        next_row = ws.max_row + 1
        
        # Add the new entry
        ws.cell(row=next_row, column=1, value=date)
        ws.cell(row=next_row, column=2, value=day)
        ws.cell(row=next_row, column=3, value=start_time)
        ws.cell(row=next_row, column=4, value=end_time)
        ws.cell(row=next_row, column=5, value=time_elapsed)
        ws.cell(row=next_row, column=6, value=task)
        
        # Save the workbook
        wb.save(FILE_PATH)
        
        print(f"Excel updated with new entry: {date} - {task}")
        return True
        
    except Exception as e:
        print(f"❌ Error updating Excel: {e}")
        import traceback
        traceback.print_exc()
        return False

# ==========================
# MAIN FUNCTIONS
# ==========================
def main():
    """Main function for the timesheet application"""
    print("\nVoice Timesheet Application\n")
    
    while True:
        command = input("\nPress Enter to record (or type 'exit' to quit, 'view' to see timesheet): ").strip().lower()
        
        if command == "exit":
            print("Exiting application. Goodbye!")
            break
            
        elif command == "view":
            display_timesheet()
            continue
        
        # Record audio
        print("\n--- Recording Audio ---")
        success = record_audio(AUDIO_FILE, duration=15)
        if not success:
            print("Failed to record audio. Please try again.")
            continue
        
        # Transcribe audio
        print("\n--- Transcribing Audio ---")
        transcript = transcribe_audio(AUDIO_FILE)
        if not transcript:
            print("Failed to transcribe audio. Please try again.")
            continue
        
        # Process with GPT
        print("\n--- Processing Transcript with GPT ---")
        extracted_data = process_with_gpt(transcript)
        if not extracted_data:
            print("Failed to extract data from transcript. Please try again.")
            continue
        
        # Parse entries and update Excel
        entries = parse_entries(extracted_data)
        if entries:
            print(f"Found {len(entries)} entries. Adding to timesheet...")
            for entry in entries:
                update_excel(**entry)
        else:
            print("No valid entries found in the transcript.")

def parse_entries(extracted_text):
    """Parse the GPT output into structured entries"""
    entries = []
    current_entry = {}
    
    for line in extracted_text.strip().split('\n'):
        line = line.strip()
        if not line:
            if current_entry and len(current_entry) >= 5:  # Ensure we have most fields
                entries.append(current_entry)
            current_entry = {}
            continue
        
        if ':' in line:
            key, value = line.split(':', 1)
            key = key.strip()
            value = value.strip()
            
            if key == "Date":
                current_entry["date"] = value
            elif key == "Day":
                current_entry["day"] = value
            elif key == "Start Time":
                current_entry["start_time"] = value
            elif key == "End Time":
                current_entry["end_time"] = value
            elif key == "Time Elapsed":
                current_entry["time_elapsed"] = value
            elif key == "Task":
                current_entry["task"] = value
    
    # Add the last entry if there is one
    if current_entry and len(current_entry) >= 5:
        entries.append(current_entry)
    
    return entries

def display_timesheet():
    """Display the current timesheet contents"""
    try:
        if not os.path.exists(FILE_PATH):
            print("No timesheet found. Record some entries first.")
            return
        
        df = pd.read_excel(FILE_PATH)
        if df.empty:
            print("Timesheet is empty. Record some entries first.")
            return
        
        print("\n📋 Current Timesheet Entries 📋")
        print("=" * 80)
        print(df.to_string(index=False))
        print("=" * 80)
        
    except Exception as e:
        print(f"Error displaying timesheet: {e}")

# ==========================
# ENTRY POINT
# ==========================
if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nApplication interrupted by user. Exiting...")
    except Exception as e:
        print(f"Unexpected error: {e}")
        import traceback
        traceback.print_exc()
